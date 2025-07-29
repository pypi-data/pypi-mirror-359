"""
Functions to read reticle calibration data, find a transformation between
coordinate frames, and apply the transformation.
"""

import csv
import io
import logging
import re
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from scipy import optimize as opt
from scipy.spatial.transform import Rotation

from aind_mri_utils.arc_angles import calculate_arc_angles
from aind_mri_utils.rotations import (
    apply_affine,
    apply_inverse_affine,
    combine_angles,
    compose_transforms,
    prepare_data_for_homogeneous_transform,
    rotation_matrix_from_vectors,
)

logger = logging.getLogger(__name__)


def _extract_calibration_metadata(ws):
    """Extract calibration metadata from an Excel worksheet.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        The worksheet object from which to extract the calibration metadata.

    Returns
    -------
    tuple
        A tuple containing:
        - global_factor (float): The global scale value.
        - global_rotation_degrees (float): The global rotation in degrees.
        - manipulator_factor (float): The manipulator scale value.
        - global_offset (numpy.ndarray): The global offset as a 3-element
          array.
        - reticle_name (str): The name of the reticle.
    """
    row_iter = ws.iter_rows(min_row=1, max_row=2, values_only=True)
    col_name_lookup = {k: i for i, k in enumerate(next(row_iter))}
    metadata_values = next(row_iter)
    global_factor = metadata_values[col_name_lookup["GlobalFactor"]]
    global_rotation_degrees = metadata_values[
        col_name_lookup["GlobalRotationDegrees"]
    ]
    manipulator_factor = metadata_values[col_name_lookup["ManipulatorFactor"]]
    reticle_name = metadata_values[col_name_lookup["Reticule"]]
    offset_x_pos = col_name_lookup["GlobalOffsetX"]
    global_offset = np.array(
        metadata_values[offset_x_pos : offset_x_pos + 3],  # noqa: E203
        dtype=float,
    )
    return (
        global_factor,
        global_rotation_degrees,
        manipulator_factor,
        global_offset,
        reticle_name,
    )


def reticle_metadata_transform(global_rotation_degrees):
    """
    Calculate the transformation matrix that will apply the global offset and
    rotation to reticle points.

    Parameters
    ----------
    global_rotation_degrees : float
        The global rotation in degrees.

    Returns
    -------
    numpy.ndarray
        The rotation matrix.
    """
    R = (
        Rotation.from_euler("z", global_rotation_degrees, degrees=True)
        .as_matrix()
        .squeeze()
    )
    return R


def _contains_none(arr):
    """
    Checks if all arguments are not None.

    Parameters
    ----------
    arr : iterable
        An iterable of elements to check.

    Returns
    -------
    bool
        True if any element is None, False otherwise.
    """
    return any(x is None for x in arr)


def _combine_pairs(list_of_pairs):
    """
    Combine lists of pairs into separate global and manipulator points
    matrices.

    Parameters
    ----------
    list_of_pairs : list of tuple
        A list of tuples, each containing a reticle point and a probe point as
        numpy arrays.

    Returns
    -------
    tuple
        Two numpy arrays, one for global points and one for manipulator points.
    """
    global_pts, manipulator_pts = [np.vstack(x) for x in zip(*list_of_pairs)]
    return global_pts, manipulator_pts


def _extract_calibration_pairs(ws):
    """
    Extract calibration pairs from an Excel worksheet.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        The worksheet object from which to extract the calibration pairs.

    Returns
    -------
    dict
        A dictionary where keys are probe names and values are lists of tuples,
        each containing a reticle point and a probe point as numpy arrays.
    """
    pair_lists_by_probe = dict()
    for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
        probe_name = row[0]
        if probe_name is None:
            continue
        reticle_pt = np.array(row[1:4])
        probe_pt = np.array(row[4:7])
        if _contains_none(reticle_pt) or _contains_none(probe_pt):
            continue
        if probe_name not in pair_lists_by_probe:
            pair_lists_by_probe[probe_name] = []
        pair_lists_by_probe[probe_name].append((reticle_pt, probe_pt))
    pair_mats_by_probe = {
        k: _combine_pairs(v) for k, v in pair_lists_by_probe.items()
    }
    return pair_mats_by_probe


def _apply_metadata_to_pair_mats(
    global_pts,
    manipulator_pts,
    global_factor,
    global_rotation_degrees,
    global_offset,
    manipulator_factor,
):
    """
    Apply calibration metadata to global and manipulator points matrices.

    Parameters
    ----------
    global_pts : numpy.ndarray
        The global points matrix.
    manipulator_pts : numpy.ndarray
        The manipulator points matrix.
    global_factor : float
        The global factor value.
    global_rotation_degrees : float
        The global rotation in degrees.
    global_offset : numpy.ndarray
        The global offset as a 3-element array.
    manipulator_factor : float
        The manipulator factor value.

    Returns
    -------
    tuple
        The adjusted global points and manipulator points matrices.
    """
    if global_rotation_degrees != 0:
        rot_mat = reticle_metadata_transform(global_rotation_degrees)
        # Transposed because points are row vectors
        global_pts = global_pts @ rot_mat.T
    global_pts = global_pts * global_factor + global_offset
    manipulator_pts = manipulator_pts * manipulator_factor
    return global_pts, manipulator_pts


def read_manual_reticle_calibration(
    filename, points_sheet_name="points", metadata_sheet_name="metadata"
):
    """Read reticle calibration data from an Excel file.

    Parameters
    ----------
    filename : str
        The path to the Excel file containing the calibration data.
    points_sheet_name : str, optional
        The name of the sheet containing the calibration points.
        The default is "points".
    metadata_sheet_name : str, optional
        The name of the sheet containing the calibration metadata.
        The default is "metadata".

    Returns
    -------
    adjusted_pairs_by_probe : dict
        Adjusted calibration pairs by probe name.
    global_offset : numpy.ndarray The global offset as a 3-element array.
    global_rotation_degrees : float
        The global rotation in degrees.
    reticle_name : str
        The name of the reticle.

    Raises
    ------
    ValueError
        If the specified sheets are not found in the Excel file.

    """
    in_mem_file = None
    with open(filename, "rb") as f:
        in_mem_file = io.BytesIO(f.read())
    wb = load_workbook(in_mem_file, read_only=True, data_only=True)
    if points_sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {points_sheet_name} not found in {filename}")
    if metadata_sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet {metadata_sheet_name} not found in {filename}"
        )
    (
        global_factor,
        global_rotation_degrees,
        manipulator_factor,
        global_offset,
        reticle_name,
    ) = _extract_calibration_metadata(wb[metadata_sheet_name])
    pairs_by_probe = _extract_calibration_pairs(wb["points"])
    adjusted_pairs_by_probe = {
        k: _apply_metadata_to_pair_mats(
            *v,
            global_factor,
            global_rotation_degrees,
            global_offset,
            manipulator_factor,
        )
        for k, v in pairs_by_probe.items()
    }
    return (
        adjusted_pairs_by_probe,
        global_offset,
        global_rotation_degrees,
        reticle_name,
    )


def read_parallax_calibration_dir(
    parallax_points_dir,
    sn_filename_regexp=re.compile(r"(?i)points_SN\d+(?:_.*)?.csv$"),
    *args,
    **kwargs,
):
    """
    Read parallax calibration data from a directory of CSV files.

    Parameters
    ----------
    parallax_points_dir : str
        The directory containing the parallax calibration CSV files.
    sn_filename_regexp : re.Pattern, optional
        The regular expression pattern to match filenames (default is
        r"(?i)points_SN\\d+(?:_.*)?.csv$").
    *args : tuple
        Additional arguments to pass to the calibration file reader.
    **kwargs : dict
        Additional keyword arguments to pass to the calibration file reader.

    Returns
    -------
    dict
        A dictionary where keys are controller numbers and values are tuples
        of numpy arrays for global and manipulator points.
    """
    pairs_by_controller = {}
    p_path = Path(parallax_points_dir)
    for filename in p_path.iterdir():
        if filename.is_file() and re.search(sn_filename_regexp, filename.name):
            _append_parallax_calibration_file(
                pairs_by_controller, filename, *args, **kwargs
            )
    mats_by_controller = {
        k: _combine_pairs(v) for k, v in pairs_by_controller.items()
    }
    return mats_by_controller


def read_parallax_calibration_file(parallax_points_filename, *args, **kwargs):
    """
    Read parallax calibration data from a single CSV file.

    Parameters
    ----------
    parallax_points_filename : str
        The path to the CSV file containing the parallax points data.
    *args : tuple
        Additional arguments to pass to the calibration file reader.
    **kwargs : dict
        Additional keyword arguments to pass to the calibration file reader.

    Returns
    -------
    dict
        A dictionary where keys are controller numbers and values are tuples
        of numpy arrays for global and manipulator points.
    """
    pairs_by_controller = {}
    _append_parallax_calibration_file(
        pairs_by_controller, parallax_points_filename, *args, **kwargs
    )
    mats_by_controller = {
        k: _combine_pairs(v) for k, v in pairs_by_controller.items()
    }
    return mats_by_controller


def read_parallax_calibration_dir_and_correct(
    parallax_calibration_dir,
    reticle_offset,
    reticle_rotation,
    local_scale_factor=1 / 1000,
    global_scale_factor=1 / 1000,
):
    """
    Read and correct parallax calibration data from a directory of CSV files.

    Parameters
    ----------
    parallax_calibration_dir : str
        Directory containing parallax calibration data.
    reticle_offset : numpy.ndarray
        Offset of the reticle.
    reticle_rotation : float
        Rotation of the reticle in degrees.
    local_scale_factor : float, optional
        Local scale factor for calibration, by default 1/1000.
    global_scale_factor : float, optional
        Global scale factor for calibration, by default 1/1000.

    Returns
    -------
    dict
        A dictionary where keys are controller numbers and values are tuples
        of numpy arrays for global and manipulator points.
    """
    pairs_by_probe = read_parallax_calibration_dir(parallax_calibration_dir)
    corrected_pairs_by_probe = {}
    for controller, pairs in pairs_by_probe.items():
        reticle_pts, manip_pts = _apply_metadata_to_pair_mats(
            *pairs,
            global_scale_factor,
            reticle_rotation,
            reticle_offset,
            local_scale_factor,
        )
        corrected_pairs_by_probe[controller] = (reticle_pts, manip_pts)
    return corrected_pairs_by_probe


def _append_parallax_calibration_file(
    pairs_by_controller,
    parallax_points_filename,
    sn_colname="sn",
    sn_regexp=re.compile(r"(\d+)$"),
):
    """
    Reads parallax calibration data from a CSV file and organizes it by
    controller number.

    Parameters
    ----------
    parallax_points_filename : str
        The path to the CSV file containing parallax points data.
    sn_colname : str, optional
        The column name for the serial number (default is "sn").
    sn_regexp : re.Pattern, optional
        The regular expression pattern to extract the controller number from
        the serial number

    Returns
    -------
    dict
        A dictionary where the keys are controller numbers and the values are
        lists of tuples. Each tuple contains two numpy arrays: the first array
        represents the reticle points (global coordinates) and the second array
        represents the manipulator points (local coordinates).

    Notes
    -----
    The CSV file is expected to have columns named "global_x", "global_y",
    "global_z" for reticle points and "local_x", "local_y", "local_z" for
    manipulator points.
    """
    dims = ["x", "y", "z"]
    reticle_colnames = [f"global_{dim}" for dim in dims]
    manipulator_colnames = [f"local_{dim}" for dim in dims]
    with open(parallax_points_filename, "r", newline="") as f:
        reader = csv.reader(f)
        header = next(reader)
        # Map each column name to its first index
        index_map = {}
        for i, col in enumerate(header):
            if col not in index_map:
                index_map[col] = i
        # Find indices for the columns we care about
        sn_col_ndx = index_map[sn_colname]
        reticle_pt_ndxs = [index_map[col] for col in reticle_colnames]
        manip_pt_ndxs = [index_map[col] for col in manipulator_colnames]
        # Read the data
        for row in reader:
            controller_no = int(re.search(sn_regexp, row[sn_col_ndx]).group(1))
            ret_pt = np.array([float(row[ndx]) for ndx in reticle_pt_ndxs])
            manip_pt = np.array([float(row[ndx]) for ndx in manip_pt_ndxs])
            # Append to this manipulator's list of pairs, creating a new list
            # if needed
            pairs_by_controller.setdefault(controller_no, []).append(
                (ret_pt, manip_pt)
            )
    return pairs_by_controller


def _unpack_theta_scale(theta):
    """
    Helper function to unpack theta into rotation matrix and translation.

    Parameters
    ----------
    theta : numpy.ndarray
        The array containing rotation angles, scale factors, and translation
        values.

    Returns
    -------
    tuple
        A tuple containing:
        - R (numpy.ndarray): The rotation matrix.
        - scale (numpy.ndarray): The scale factors.
        - translation (numpy.ndarray): The translation vector.
    """
    R = combine_angles(*theta[0:3])
    scale = theta[3:6]
    translation = theta[6:]
    return R, scale, translation


def _fit_params_with_scaling(reticle_pts, probe_pts, lamb=0.1, **kwargs):
    """
    Fit rotation parameters to align reticle points with probe points using
    least squares optimization. The rotation matrix and translation vector
    are the solution for the equation

    probe_pts = R @ reticle_pts + translation

    where each point is a column vector.

    Because numpy is row-major, points are often stored as row vectors. In this
    case, you should use the transpose of this equation:

    probe_pts = reticle_pts @ R.T + translation

    Parameters
    ----------
    reticle_pts : numpy.ndarray
        The reticle points to be transformed.
    probe_pts : numpy.ndarray
        The probe points to align with.
    find_scaling : bool, optional
        If True, find a scaling factor to apply to the reticle points.
        The default is True.
    **kwargs : dict
        Additional keyword arguments to pass to the least squares optimization
        function.

    Returns
    -------
    tuple
        A tuple containing:
        - R (numpy.ndarray): The 3x3 rotation matrix.
        - translation (numpy.ndarray): The 3-element translation vector.
        - scaling (float): The scaling factor.
    """

    def residuals(theta, probe_pts, bregma_pts, lamb):
        """cost function for least squares optimization"""
        R, scale, translation = _unpack_theta_scale(theta)
        transformed_reticle = bregma_pts @ R.T * scale + translation
        res = (transformed_reticle - probe_pts).flatten()
        if lamb == 0:
            return res
        else:
            reg = np.sqrt(lamb) * (np.abs(scale) - np.ones(3))
            return np.concatenate((res, reg))

    # Initial guess of parameters
    theta0 = np.zeros(9)
    theta0[3:6] = 1.0
    npt = probe_pts.shape[0]
    if npt > 1:
        # Initial guess of rotation: align the vectors between the first
        # two points
        for other_pt in range(1, npt):
            probe_diff = probe_pts[other_pt, :] - probe_pts[0, :]
            reticle_diff = reticle_pts[other_pt, :] - reticle_pts[0, :]
            reticle_norm = np.linalg.norm(reticle_diff.squeeze())
            if reticle_norm > 0:
                break
        if reticle_norm > 0:
            R_init = rotation_matrix_from_vectors(
                reticle_diff.squeeze(), probe_diff.squeeze()
            )
            theta0[0:3] = Rotation.from_matrix(R_init).as_euler("xyz")

    # Initial guess of translation: find the point on the reticle closest to
    # zero
    smallest_pt = np.argmin(np.linalg.norm(reticle_pts, axis=1))
    theta0[6:] = probe_pts[smallest_pt, :]

    res = opt.least_squares(
        residuals,
        theta0,
        args=(probe_pts, reticle_pts, lamb),
        **kwargs,
    )
    R, scale, translation = _unpack_theta_scale(res.x)
    scale_positive = np.abs(scale)
    scale_sign = np.sign(scale)
    R = np.diag(scale_sign) @ R
    R = np.diag(scale_positive) @ R
    return R, translation, scale_positive


def _unpack_theta(theta):
    """
    Helper function to unpack theta into rotation matrix and translation.

    Parameters
    ----------
    theta : numpy.ndarray
        The array containing rotation angles and translation values.

    Returns
    -------
    tuple
        A tuple containing:
        - R (numpy.ndarray): The rotation matrix.
        - offset (numpy.ndarray): The translation vector.
    """
    R = combine_angles(*theta[0:3])
    offset = theta[3:6]
    return R, offset


def fit_rotation_params(bregma_pts, probe_pts, find_scaling=True, **kwargs):
    """
    Fit rotation parameters to align bregma points with probe points using
    least squares optimization. The rotation matrix and translation vector
    are the solution for the equation

    probe_pts = R @ bregma_pts + translation

    where each point is a column vector.

    Because numpy is row-major, points are often stored as row vectors. In this
    case, you should use the transpose of this equation:

    probe_pts = bregma_pts @ R.T + translation

    Parameters
    ----------
    bregma_pts : numpy.ndarray
        The bregma points to be transformed.
    probe_pts : numpy.ndarray
        The probe points to align with.
    **kwargs : dict
        Additional keyword arguments to pass to the least squares optimization
        function.

    Returns
    -------
    tuple
        A tuple containing:
        - R (numpy.ndarray): The 3x3 rotation matrix.
        - translation (numpy.ndarray): The 3-element translation vector.
    """
    if bregma_pts.shape != probe_pts.shape:
        raise ValueError("bregma_pts and probe_pts must have the same shape")
    if bregma_pts.shape[1] != 3:
        raise ValueError("bregma_pts and probe_pts must have 3 columns")

    if find_scaling:
        return _fit_params_with_scaling(bregma_pts, probe_pts, **kwargs)

    R_homog = np.eye(4)
    bregma_pts_homog = prepare_data_for_homogeneous_transform(bregma_pts)
    transformed_pts_homog = np.empty_like(bregma_pts_homog)

    def fun(theta):
        """cost function for least squares optimization"""
        R_homog[0:3, 0:3] = combine_angles(*theta[0:3])
        R_homog[0:3, 3] = theta[3:6]  # translation
        np.matmul(bregma_pts_homog, R_homog.T, out=transformed_pts_homog)
        residuals = (transformed_pts_homog[:, 0:3] - probe_pts).flatten()
        return residuals

    # Initial guess of parameters
    theta0 = np.zeros(6)

    if probe_pts.shape[0] > 1:
        # Initial guess of rotation: align the vectors between the first
        # two points
        probe_diff = np.diff(probe_pts[:2, :], axis=0)
        bregma_diff = np.diff(bregma_pts[:2, :], axis=0)
        R_init = rotation_matrix_from_vectors(
            bregma_diff.squeeze(), probe_diff.squeeze()
        )
        theta0[0:3] = Rotation.from_matrix(R_init).as_euler("xyz")

    # Initial guess of translation: find the point on the bregma closest to
    # zero
    smallest_pt = np.argmin(np.linalg.norm(bregma_pts, axis=1))
    theta0[3:6] = probe_pts[smallest_pt, :]

    res = opt.least_squares(fun, theta0, **kwargs)
    R, translation = _unpack_theta(res.x)
    scaling = np.ones(3)
    return R, translation, scaling


def _fit_by_probe(pairs_by_probe, *args, **kwargs):
    """
    Fit rotation parameters for each probe.

    Parameters
    ----------
    pairs_by_probe : dict
        A dictionary where keys are probe names and values are tuples of numpy
        arrays for reticle and probe points.
    *args : tuple
        Additional arguments to pass to the fitting function.
    **kwargs : dict
        Additional keyword arguments to pass to the fitting function.

    Returns
    -------
    dict
        A dictionary where keys are probe names and values are tuples
        containing the rotation matrix, translation vector, and scaling factor.
    """
    cal_by_probe = {
        k: fit_rotation_params(*v, *args, **kwargs)
        for k, v in pairs_by_probe.items()
    }
    return cal_by_probe


def fit_rotation_params_from_manual_calibration(filename, *args, **kwargs):
    """
    Fits rotation parameters from manual calibration data.

    Parameters
    ----------
    filename : str
        File containing the calibration data, stored as a `.xslx` file.
    *args : tuple
        Additional arguments to pass to the fitting function. See
        `fit_rotation_params`.
    **kwargs : dict
        Additional keyword arguments to pass to the fitting function.
        See `fit_rotation_params`.

    Returns
    -------
    cal_by_probe : dict
        Calibration parameters by probe. Each value is a tuple containing the
        rotation matrix, translation vector, and scaling factor. These
        parameters transform bregma points to probe points, both in
        millimeters.
    R_reticle_to_bregma : numpy.ndarray
        Rotation matrix from reticle to bregma coordinates, in millimeters.
    t_reticle_to_bregma : float
        Translation vector from reticle to bregma coordinates, in millimeters.
    """
    adjusted_pairs_by_probe, global_offset, global_rotation_degrees, _ = (
        read_manual_reticle_calibration(filename)
    )
    R_reticle_to_bregma = reticle_metadata_transform(global_rotation_degrees)
    cal_by_probe = _fit_by_probe(adjusted_pairs_by_probe, *args, **kwargs)
    return cal_by_probe, R_reticle_to_bregma, global_offset


def fit_rotation_params_from_parallax(
    parallax_calibration_dir,
    reticle_offset,
    reticle_rotation,
    local_scale_factor=1 / 1000,
    global_scale_factor=1 / 1000,
    *args,
    **kwargs,
):
    """
    Fits rotation parameters from parallax calibration data.

    Parameters
    ----------
    parallax_calibration_dir : str
        Directory containing parallax calibration data.
    reticle_offset : float
        Offset of the reticle, i.e. the location of the center of the reticle
        pattern in bregma-relative coordinates.
    reticle_rotation : float
        Rotation of the reticle. i.e. the angle about the z-axis of the reticle
        coordinate system and the bregma coordinate system in degrees. The sign
        of the angle should be positive if the reticle is rotated clockwise
        w.r.t bregma coordinate system when viewed from above.
    local_scale_factor : float, optional
        Local scale factor for calibration, by default 1/1000. Parallax stores
        points in microns, so this factor is used to convert to millimeters.
    global_scale_factor : float, optional
        Global scale factor for calibration, by default 1/1000. Parallax stores
        points in microns, so this factor is used to convert to millimeters.
    *args : tuple
        Additional arguments to pass to the fitting function. See
        `fit_rotation_params`.
    **kwargs : dict
        Additional keyword arguments to pass to the fitting function.
        See `fit_rotation_params`.

    Returns
    -------
    cal_by_probe : dict
        Calibration parameters by probe. Each value is a tuple containing the
        rotation matrix, translation vector, and scaling factor. These
        parameters transform bregma points to probe points, both in
        millimeters.
    R_reticle_to_bregma : numpy.ndarray
        Rotation matrix from reticle to bregma coordinates, in millimeters.
    """
    adjusted_pairs_by_probe = read_parallax_calibration_dir_and_correct(
        parallax_calibration_dir,
        reticle_offset,
        reticle_rotation,
        local_scale_factor,
        global_scale_factor,
    )
    cal_by_probe = _fit_by_probe(adjusted_pairs_by_probe, *args, **kwargs)
    R_reticle_to_bregma = reticle_metadata_transform(reticle_rotation)
    return cal_by_probe, R_reticle_to_bregma


def _debug_print_pt_err(reticle, probe, predicted_probe, err, decimals=3):
    """
    Print the error for a single point.

    Parameters
    ----------
    reticle : numpy.ndarray
        The reticle point.
    probe : numpy.ndarray
        The probe point.
    predicted_probe : numpy.ndarray
        The predicted probe point.
    err : float
        The error value.
    decimals : int, optional
        The number of decimal places to round to (default is 3).
    """
    rounded_reticle = np.round(reticle, decimals=decimals)
    rounded_probe = np.round(probe, decimals=decimals)
    rounded_pred = np.round(predicted_probe, decimals=decimals)
    logger.debug(
        f"Reticle {rounded_reticle} -> "
        f"Probe {rounded_probe}: predicted {rounded_pred} "
        f"error {err:.2f} µm"
    )


def _debug_print_err_stats(errs):
    """
    Print error statistics for a probe.

    Parameters
    ----------
    name : str
        The name of the probe.
    errs : numpy.ndarray
        The array of error values.
    """
    logger.debug(
        f"mean error {errs.mean():.2f} µm, max error {errs.max():.2f} µm"
    )


def _debug_fits(
    cal_by_probe,
    R_reticle_to_bregma,
    t_reticle_to_bregma,
    adjusted_pairs_by_probe,
):
    """
    Debug the fits for each probe.

    Parameters
    ----------
    cal_by_probe : dict
        Calibration parameters by probe.
    R_reticle_to_bregma : numpy.ndarray
        Rotation matrix from reticle to bregma coordinates.
    t_reticle_to_bregma : numpy.ndarray
        Translation vector from reticle to bregma coordinates.
    adjusted_pairs_by_probe : dict
        Adjusted calibration pairs by probe.

    Returns
    -------
    dict
        A dictionary where keys are probe names and values are arrays of error
        values.
    """
    errs_by_probe = {}
    for probe, (bregma_pts, probe_pts) in adjusted_pairs_by_probe.items():
        R, t, scaling = cal_by_probe[probe]
        predicted_probe_pts = transform_bregma_to_probe(bregma_pts, R, t)
        # in mm
        errs = np.linalg.norm(predicted_probe_pts - probe_pts, axis=1)
        errs_by_probe[probe] = errs
        if logger.isEnabledFor(logging.DEBUG):
            logger.debug(f"Probe {probe}:")
            logger.debug("rotation:")
            logger.debug(R)
            logger.debug(f"translation: {t}")
            logger.debug(f"scaling: {scaling}")
            _debug_print_err_stats(1000 * errs)
            reticle_pts = transform_probe_to_bregma(
                bregma_pts, R_reticle_to_bregma, t_reticle_to_bregma
            )
            for i in range(len(errs)):
                _debug_print_pt_err(
                    reticle_pts[i],
                    probe_pts[i],
                    predicted_probe_pts[i],
                    1000 * errs[i],
                )
    return errs_by_probe


def debug_manual_calibration(filename, *args, **kwargs):
    """
    Debugs the manual calibration process by fitting rotation parameters and
    reading manual reticle calibration data.

    Parameters
    ----------
    filename : str
        The path to the file containing manual calibration data.
    *args : tuple
        Additional positional arguments to be passed to the calibration
        functions.
    **kwargs : dict
        Additional keyword arguments to be passed to the calibration functions.

    Returns
    -------
    cal_by_probe : dict
        Calibration data for each probe.
    R_reticle_to_bregma : numpy.ndarray
        Rotation matrix from reticle to bregma.
    t_reticle_to_bregma : numpy.ndarray
        Translation vector from reticle to bregma.
    adjusted_pairs_by_probe : dict
        Adjusted pairs of calibration data by probe.
    errs_by_probe : dict
        Errors in the fits for each probe.
    """
    cal_by_probe, R_reticle_to_bregma, t_reticle_to_bregma = (
        fit_rotation_params_from_manual_calibration(filename, *args, **kwargs)
    )
    (
        adjusted_pairs_by_probe,
        global_offset,
        global_rotation_degrees,
        _,
    ) = read_manual_reticle_calibration(filename)
    errs_by_probe = _debug_fits(
        cal_by_probe,
        R_reticle_to_bregma,
        t_reticle_to_bregma,
        adjusted_pairs_by_probe,
    )
    return (
        cal_by_probe,
        R_reticle_to_bregma,
        t_reticle_to_bregma,
        adjusted_pairs_by_probe,
        errs_by_probe,
    )


def debug_parallax_calibration(
    parallax_calibration_dir,
    reticle_offset,
    reticle_rotation,
    local_scale_factor=1 / 1000,
    global_scale_factor=1 / 1000,
    *args,
    **kwargs,
):
    """
    Debugs the parallax calibration process by reading calibration data,
    applying corrections, fitting the data, and calculating errors.

    Parameters
    ----------
    parallax_calibration_dir : str
        Directory containing the parallax calibration data.
    reticle_offset : array-like
        Offset of the reticle in the calibration setup.
    reticle_rotation : array-like
        Rotation of the reticle in the calibration setup.
    local_scale_factor : float, optional
        Scale factor for local adjustments, by default 1/1000.
    global_scale_factor : float, optional
        Scale factor for global adjustments, by default 1/1000.
    *args : tuple
        Additional arguments to pass to the fitting function.
    **kwargs : dict
        Additional keyword arguments to pass to the fitting function.

    Returns
    -------
    cal_by_probe : dict
        Calibration data organized by probe.
    R_reticle_to_bregma : array-like
        Transformation matrix from reticle coordinates to bregma coordinates.
    adjusted_pairs_by_probe : dict
        Adjusted calibration pairs organized by probe.
    errs_by_probe : dict
        Errors in the calibration fits organized by probe.
    """
    adjusted_pairs_by_probe = read_parallax_calibration_dir_and_correct(
        parallax_calibration_dir,
        reticle_offset,
        reticle_rotation,
        local_scale_factor,
        global_scale_factor,
    )
    cal_by_probe = _fit_by_probe(adjusted_pairs_by_probe, *args, **kwargs)
    R_reticle_to_bregma = reticle_metadata_transform(reticle_rotation)
    errs_by_probe = _debug_fits(
        cal_by_probe,
        R_reticle_to_bregma,
        reticle_offset,
        adjusted_pairs_by_probe,
    )
    return (
        cal_by_probe,
        R_reticle_to_bregma,
        adjusted_pairs_by_probe,
        errs_by_probe,
    )


def transform_bregma_to_probe(bregma_pts, R, translation):
    """
    Transform reticle points to probe points using rotation and translation.

    Parameters
    ----------
    bregma_pts : np.array(N,3)
        Bregma points to transform.
    R : np.array(3,3)
        Affine matrix as provided from fit functions in this module.
    translation : np.array(3,)
        Translation vector as provided from fit functions in this module.

    Returns
    -------
    np.array(N,3)
        Transformed probe points.

    Notes
    -----
    Expects the affine and transform to take bregma points to probe points.
    """
    return apply_affine(bregma_pts, R, translation)


def transform_probe_to_bregma(probe_pts, R, translation):
    """
    Transform probe points to reticle points using rotation and translation.

    Parameters
    ----------
    probe_pts : np.array(N,3)
        Probe points to transform.
    R : np.array(3,3)
        Affine matrix as provided from fit functions in this module.
    translation : np.array(3,)
        Translation vector as provided from fit functions in this module.

    Returns
    -------
    np.array(N,3)
        Transformed bregma points.

    Notes
    -----
    Expects the affine and transform to take bregma points to probe points.
    """
    return apply_inverse_affine(probe_pts, R, translation)


def transform_bregma_to_reticle(bregma_pts, R, translation):
    """
    Transform bregma points to reticle points using rotation and translation.

    Parameters
    ----------
    bregma_pts : numpy.ndarray
        The bregma points to transform.
    R : np.array(3,3)
        Affine matrix as provided from fit functions in this module.
    translation : np.array(3,)
        Translation vector as provided from fit functions in this module.

    Returns
    -------
    numpy.ndarray
        The transformed reticle points.
    """
    return apply_inverse_affine(bregma_pts, R, translation)


def transform_reticle_to_bregma(reticle_pts, R, translation):
    """
    Transform reticle points to bregma points using rotation and translation.

    Parameters
    ----------
    reticle_pts : numpy.ndarray
        The reticle points to transform.
    R : np.array(3,3)
        Affine matrix as provided from fit functions in this module.
    translation : np.array(3,)
        Translation vector as provided from fit functions in this module.

    Returns
    -------
    numpy.ndarray
        The transformed bregma points.
    """
    return apply_affine(reticle_pts, R, translation)


def combine_reticle_to_probe_transforms(
    R_bregma_to_probe,
    t_bregma_to_probe,
    R_reticle_to_bregma,
    t_reticle_to_bregma,
):
    """
    Combines the transformation matrices and translation vectors from reticle
    to bregma and bregma to probe.

    Parameters
    ----------
    R_bregma_to_probe : numpy.ndarray
        Rotation matrix from bregma to probe.
    t_bregma_to_probe : numpy.ndarray
        Translation vector from bregma to probe.
    R_reticle_to_bregma : numpy.ndarray
        Rotation matrix from reticle to bregma.
    t_reticle_to_bregma : numpy.ndarray
        Translation vector from reticle to bregma.

    Returns
    -------
    numpy.ndarray
        Combined rotation matrix and translation vector from reticle to probe.
    """
    return compose_transforms(
        R_reticle_to_bregma,
        t_reticle_to_bregma,
        R_bregma_to_probe,
        t_bregma_to_probe,
    )


def transform_reticle_to_probe(
    reticle_pts,
    R_bregma_to_probe,
    t_bregma_to_probe,
    R_reticle_to_bregma,
    t_reticle_to_bregma,
):
    """
    Transform reticle points to probe points using rotation and translation.

    Parameters
    ----------
    reticle_pts : numpy.ndarray
        The reticle points to transform.
    R_bregma_to_probe : numpy.ndarray
        Rotation matrix from bregma to probe coordinates.
    t_bregma_to_probe : numpy.ndarray
        Translation vector from bregma to probe coordinates.
    R_reticle_to_bregma : numpy.ndarray
        Rotation matrix from reticle to bregma coordinates.
    t_reticle_to_bregma : numpy.ndarray
        Translation vector from reticle to bregma coordinates.

    Returns
    -------
    numpy.ndarray
        The transformed probe points.
    """
    R_reticle_to_probe, t_reticle_to_probe = (
        combine_reticle_to_probe_transforms(
            R_reticle_to_bregma,
            t_reticle_to_bregma,
            R_bregma_to_probe,
            t_bregma_to_probe,
        )
    )
    return apply_affine(reticle_pts, R_reticle_to_probe, t_reticle_to_probe)


def transform_probe_to_reticle(
    probe_pts,
    R_bregma_to_probe,
    t_bregma_to_probe,
    R_reticle_to_bregma,
    t_reticle_to_bregma,
):
    """
    Transforms probe points to reticle coordinates.

    Parameters
    ----------
    probe_pts : ndarray
        Array of points in probe coordinates.
    R_bregma_to_probe : ndarray
        Rotation matrix from bregma to probe coordinates.
    t_bregma_to_probe : ndarray
        Translation vector from bregma to probe coordinates.
    R_reticle_to_bregma : ndarray
        Rotation matrix from reticle to bregma coordinates.
    t_reticle_to_bregma : ndarray
        Translation vector from reticle to bregma coordinates.

    Returns
    -------
    ndarray
        Transformed points in reticle coordinates.
    """
    R_reticle_to_probe, t_reticle_to_probe = (
        combine_reticle_to_probe_transforms(
            R_reticle_to_bregma,
            t_reticle_to_bregma,
            R_bregma_to_probe,
            t_bregma_to_probe,
        )
    )
    return apply_inverse_affine(
        probe_pts, R_reticle_to_probe, t_reticle_to_probe
    )


def find_probe_insertion_vector(R, newscale_z_down=np.array([0, 0, 1])):
    """
    Find the probe insertion vector from the rotation matrix and translation
    vector.

    Parameters
    ----------
    R : numpy.ndarray
        The 3x3 rotation matrix.
    translation : numpy.ndarray
        The 3-element translation vector.

    Returns
    -------
    numpy.ndarray
        The probe insertion vector.
    """
    # Probe coordinate system has z-axis pointing down
    z_axis = transform_probe_to_bregma(newscale_z_down, R, np.zeros(3))
    return z_axis


def find_probe_angle(R, newscale_z_down=np.array([0, 0, 1]), **kwargs):
    """
    Find the probe angle from the calibration rotation matrix.

    Parameters
    ----------
    R : numpy.ndarray
        The 3x3 rotation matrix.
    translation : numpy.ndarray
        The 3-element translation vector.
    newscale_z_down : numpy.ndarray, optional
        The new z-axis pointing down vector (default is [0, 0, 1]).
    **kwargs : dict
        Additional keyword arguments to pass to the `calculate_arc_angles`

    Returns
    -------
    tuple of float
        The calculated arc angles in degrees. The first element is the angle
        around the x-axis, and the second element is the angle around the
        y-axis.  Returns None if the input vector is a zero vector.
    """
    # Probe coordinate system has z-axis pointing down
    z_axis = find_probe_insertion_vector(R, newscale_z_down=newscale_z_down)
    return calculate_arc_angles(z_axis, **kwargs)


def _validate_combined_calibration_inputs(
    manual_calibration_files, parallax_directories
):
    if isinstance(manual_calibration_files, list):
        if len(manual_calibration_files) == 0:
            raise ValueError("No manual calibration files provided")
    else:
        manual_calibration_files = [manual_calibration_files]
    if not isinstance(parallax_directories, list):
        parallax_directories = [parallax_directories]
    return manual_calibration_files, parallax_directories


def combine_parallax_and_manual_calibrations(
    manual_calibration_files,
    parallax_directories,
    probes_to_ignore_manual=[],
    find_scaling=None,
    find_scaling_parallax=None,
    find_scaling_manual=None,
    *args,
    **kwargs,
):
    """Combines parallax and manual calibration data

    In each list of calibrations, calibrations later in the list will take
    priority over the earlier ones. Manual calibrations will take priority over
    parallax calibrations, unless specified by `probe_to_ignore_manual`.

    Parameters
    ----------
    manual_calibration_files : list of str
        List of files containing manual calibration data.
    parallax_directories : list of str
        List of directories containing parallax calibration data.
    probes_to_ignore_manual : list of str, optional
        List of probe names to ignore from the manual calibrations, by default
        [].
    find_scaling : bool, optional
        If True, find a scaling factor to apply to the calibration data, by
        default None, which means the default behavior in the fit function is
        used.
    find_scaling_parallax : bool, optional
        If True, find a scaling factor to apply to the parallax calibration
        data, by default None, which means the default behavior in the fit
        function is used. Overrides `find_scaling` if not None.
    find_scaling_manual : bool, optional
        If True, find a scaling factor to apply to the manual calibration data,
        by default None, which means the default behavior in the fit function
        is used. Overrides `find_scaling` if not None.
    *args : tuple
        Additional positional arguments to pass to the fitting functions.
    **kwargs : dict
        Additional keyword arguments to pass to the fitting functions.

    Returns
    -------
    cal_by_probe_combined : dict
        Combined calibration data by probe.
    R_reticle_to_bregma : numpy.ndarray
        Rotation matrix from reticle to bregma.
    global_offset : numpy.ndarray
        Global offset applied to the calibration data.

    Raises
    ------
    ValueError
        If no manual calibration files are provided.
    ValueError
        If no parallax directories are provided.

    """
    manual_calibration_files, parallax_directories = (
        _validate_combined_calibration_inputs(
            manual_calibration_files, parallax_directories
        )
    )
    # Read the first manual calibration to get the reticle metadata
    first_manual = manual_calibration_files[0]
    adjusted_pairs_by_probe, global_offset, global_rotation_degrees, _ = (
        read_manual_reticle_calibration(first_manual)
    )
    R_reticle_to_bregma = reticle_metadata_transform(global_rotation_degrees)

    # Fit the manual calibrations
    manual_scaling_kwargs = {
        "find_scaling": v
        for v in [find_scaling, find_scaling_manual]
        if v is not None
    }
    cal_by_probe_manual = _fit_by_probe(
        adjusted_pairs_by_probe, *args, **kwargs
    )
    comb_kwargs = {**manual_scaling_kwargs, **kwargs}
    for manual_calibration_file in manual_calibration_files[1:]:
        cal_by_probe, _, _ = fit_rotation_params_from_manual_calibration(
            manual_calibration_file, *args, **comb_kwargs
        )
        cal_by_probe_manual.update(cal_by_probe)

    # Fit the parallax calibrations
    parallax_scaling_kwargs = {
        "find_scaling": v
        for v in [find_scaling, find_scaling_parallax]
        if v is not None
    }
    comb_kwargs = {**parallax_scaling_kwargs, **kwargs}
    cal_by_probe_combined = {}
    for parallax_dir in parallax_directories:
        cal_by_probe, _ = fit_rotation_params_from_parallax(
            parallax_dir,
            global_offset,
            global_rotation_degrees,
            *args,
            **comb_kwargs,
        )
        cal_by_probe_combined.update(cal_by_probe)

    # Drop any probes from exclusion list from the manual calibrations
    for probe_name in probes_to_ignore_manual:
        cal_by_probe_manual.pop(probe_name, None)
    # Add the first manual calibration to the combined calibrations
    # Because they are added last, these manual calibrations take priority
    cal_by_probe_combined.update(cal_by_probe_manual)

    return cal_by_probe_combined, R_reticle_to_bregma, global_offset


def debug_parallax_and_manual_calibrations(
    manual_calibration_files,
    parallax_directories,
    probes_to_ignore_manual=[],
    find_scaling=None,
    find_scaling_parallax=None,
    find_scaling_manual=None,
    local_scale_factor=1 / 1000,
    global_scale_factor=1 / 1000,
    *args,
    **kwargs,
):
    """Debugs combined parallax and manual calibrations

    In each list of calibrations, calibrations later in the list will take
    priority over the earlier ones. Manual calibrations will take priority over
    parallax calibrations, unless specified by `probe_to_ignore_manual`.

    Parameters
    ----------
    manual_calibration_files : list of str
        List of file paths to the manual calibration files.
    parallax_directories : list of str
        List of directories containing parallax calibration data.
    probes_to_ignore_manual : list of str, optional
        List of probe names to ignore from the manual calibration data, by
        default [].
    find_scaling : bool, optional
        If True, find a scaling factor to apply to the calibration data, by
        default None, which means the default behavior in the fit function is
        used.
    find_scaling_parallax : bool, optional
        If True, find a scaling factor to apply to the parallax calibration
        data, by default None, which means the default behavior in the fit
        function is used. Overrides `find_scaling` if not None.
    find_scaling_manual : bool, optional
        If True, find a scaling factor to apply to the manual calibration data,
        by default None, which means the default behavior in the fit function
        is used. Overrides `find_scaling` if not None.
    local_scale_factor : float, optional
        Local scale factor to apply to the calibration data, by default
        1 / 1000.
    global_scale_factor : float, optional
        Global scale factor to apply to the calibration data, by default
        1 / 1000.
    *args : tuple
        Additional positional arguments to pass to the calibration functions.
    **kwargs : dict
        Additional keyword arguments to pass to the calibration functions.

    Returns
    -------
    combined_cal_by_probe : dict
        Combined calibration data by probe.
    R_reticle_to_bregma : numpy.ndarray
        Rotation matrix from reticle to bregma.
    t_reticle_to_bregma : numpy.ndarray
        Translation vector from reticle to bregma.
    combined_pairs_by_probe : dict
        Combined pairs of calibration data by probe.
    errs_by_probe : dict
        Errors by probe from the debug fits.
    """
    manual_calibration_files, parallax_directories = (
        _validate_combined_calibration_inputs(
            manual_calibration_files, parallax_directories
        )
    )
    # Read the first manual calibration to get the reticle metadata

    # Fit the manual calibrations
    manual_cal_by_probe = {}
    manual_pairs_by_probe = {}

    manual_scaling_kwargs = {
        "find_scaling": v
        for v in [find_scaling, find_scaling_manual]
        if v is not None
    }
    comb_kwargs = {**manual_scaling_kwargs, **kwargs}
    for filename in manual_calibration_files:
        cal_by_probe, R_reticle_to_bregma, t_reticle_to_bregma = (
            fit_rotation_params_from_manual_calibration(
                filename, *args, **comb_kwargs
            )
        )
        manual_cal_by_probe.update(cal_by_probe)
        (
            adjusted_pairs_by_probe,
            global_offset,
            global_rotation_degrees,
            _,
        ) = read_manual_reticle_calibration(filename)
        manual_pairs_by_probe.update(adjusted_pairs_by_probe)

    parallax_scaling_kwargs = {
        "find_scaling": v
        for v in [find_scaling, find_scaling_parallax]
        if v is not None
    }
    comb_kwargs = {**parallax_scaling_kwargs, **kwargs}
    combined_cal_by_probe = {}
    combined_pairs_by_probe = {}
    for parallax_dir in parallax_directories:
        adjusted_pairs_by_probe = read_parallax_calibration_dir_and_correct(
            parallax_dir,
            global_offset,
            global_rotation_degrees,
            local_scale_factor,
            global_scale_factor,
        )
        combined_cal_by_probe.update(
            _fit_by_probe(adjusted_pairs_by_probe, *args, **comb_kwargs)
        )
        combined_pairs_by_probe.update(adjusted_pairs_by_probe)
    for probe_name in probes_to_ignore_manual:
        manual_cal_by_probe.pop(probe_name, None)
        manual_pairs_by_probe.pop(probe_name, None)
    combined_cal_by_probe.update(manual_cal_by_probe)
    combined_pairs_by_probe.update(manual_pairs_by_probe)
    errs_by_probe = _debug_fits(
        combined_cal_by_probe,
        R_reticle_to_bregma,
        t_reticle_to_bregma,
        combined_pairs_by_probe,
    )
    return (
        combined_cal_by_probe,
        R_reticle_to_bregma,
        t_reticle_to_bregma,
        combined_pairs_by_probe,
        errs_by_probe,
    )
