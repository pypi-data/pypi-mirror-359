from abc import ABC, abstractmethod
from typing import Dict
import csv, sqlite3
import requests
import os

from openpyxl import load_workbook
import pandas as pd
from requests.models import Response


class Datasource(ABC):
    """Base class for all data sources"""
    @abstractmethod
    def extract_data(self, **kwargs):
        """Method to extract data from datasources"""
        raise NotImplementedError("Subclasses must implement this method")


class CSVSource(Datasource):
    """CSV file data source"""
    def __init__(self, file_path: str=None, url: str=None):
        self.file_path = file_path
        self.url = url
        
    def _download_csv(self, url: str):
        response = requests.get(url)

        if response.status_code == 200:
            local_file_path = os.path.join(os.getcwd(), 'downloaded_data.csv')
            with open(local_file_path, 'wb') as file:
                file.write(response.content)
                print("Downloaded CSV file Successfully📎")
                return local_file_path
        else:
            raise Exception(f"Failed to download file. Status code: {response.status_code}")

    def extract_data(self) -> pd.DataFrame:
        if self.url is not None:
            file_path = self._download_csv(self.url)
        elif self.file_path is not None:
            file_path = self.file_path
        else:
            raise ValueError("Either file_path or url must be provided")
        data = []
        with open(file_path, 'r') as file:
            reader = csv.DictReader(file)
            for row in reader:
                data.append(row)
        data = pd.DataFrame(data)
        return data

class SQLiteSource(Datasource):
    """sqlite file data source"""
    def __init__(self, connection_string: str, query: str):
        self.connection_string = connection_string
        self.query = query

    def extract_data(self) -> pd.DataFrame:
        with sqlite3.connect(self.connection_string) as conn:
            data = pd.read_sql_query(self.query, conn)
        return data.to_dict('records')

class XLSsource(Datasource):
    """Excel worksheet datasource"""
    def __init__(self, file_path: str=None, sheet_name: str=None, *args):
        self.file_path = file_path
        self.sheet_name = sheet_name

    def extract_data(self) -> pd.DataFrame:
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"File {self.file_path} does not exist")
        workbook = load_workbook(self.file_path)
        sheet = workbook[self.sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        workbook.close()
        data = pd.DataFrame(data)




