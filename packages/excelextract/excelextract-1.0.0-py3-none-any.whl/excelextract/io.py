
import os
import csv
import glob
import warnings
import sys
import io

from openpyxl import load_workbook
from openpyxl import Workbook

from .logger import logger
from .extract import extract
from .type import detectTypeOfList, convertRowToType

def loopFiles(exportConfig):
    if "input" not in exportConfig:
        raise ValueError("Missing 'input' in exportConfig")
    inputGlobs = exportConfig["input"]
    if not isinstance(inputGlobs, list):
        inputGlobs = [inputGlobs]

    inputFiles = []

    for inputGlob in inputGlobs:
        if not isinstance(inputGlob, str):
            raise ValueError(f"Invalid input glob: {inputGlob}")
        inputFiles.extend(glob.glob(inputGlob, recursive=True))

    if not inputFiles:
        raise ValueError(f"No files found matching input glob(s): {inputGlobs}")       
        
    allRows = []
    allTypes = {}

    for inputFile in inputFiles:
        if not os.path.isfile(inputFile):
            raise ValueError(f"Input file not found: {inputFile}")
        if not inputFile.endswith(".xlsx"):
            raise ValueError(f"Input file is not an Excel file: {inputFile}")
        inputFileName = os.path.basename(inputFile)
        if inputFileName.startswith("~$"):
            continue
        
        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", category=UserWarning)
                wb = load_workbook(inputFile, data_only=True)
        except Exception as e:
            raise ValueError(f"Error opening file {inputFile}: {e}")
        
        rows, types = extract(exportConfig, wb, inputFile)

        logger.info(f"Processing {inputFile} with {len(rows)} rows extracted.")

        allRows.extend(rows)

        for key, value in types.items():
            if key not in allTypes:
                allTypes[key] = value
            else:
                if allTypes[key] != value:
                    raise ValueError(f"Column type mismatch for column '{key}' in file '{inputFile}': {allTypes[key]} vs {value}")

    if len(allRows) == 0:
        raise ValueError("No rows extracted from the input files")
 
    colNames = allTypes.keys()

    if "order" in exportConfig:
        order = exportConfig["order"]
        if not isinstance(order, list):
            raise ValueError(f"Invalid order: {order}")
        colNames = [col for col in order if col in allTypes] + [col for col in allTypes if col not in order]
    
    for colName in colNames:
        if allTypes[colName] == "auto":
            data = [d[colName] for d in allRows if colName in d]
            allTypes[colName] = detectTypeOfList(data)

    out = None
    csvFileName = None
    xlsxFileName = None

    if "output" in exportConfig:
        outputFile = exportConfig["output"]

        basename = os.path.basename(outputFile)
        if "|" in basename and "." in basename:
            extensionGroup = outputFile.split(".")[-1]
            if "csv" in extensionGroup.lower():
                csvFileName = outputFile.split(".")[0] + ".csv"
            if "xlsx" in extensionGroup.lower():
                xlsxFileName = outputFile.split(".")[0] + ".xlsx"
        else:
            if str(outputFile).endswith(".csv"):
                csvFileName = outputFile
            elif str(outputFile).endswith(".xlsx"):
                xlsxFileName = outputFile
            else:
                csvFileName = outputFile + ".csv"

        outputDir = os.path.dirname(outputFile)
        if outputDir != "" and not os.path.exists(outputDir):
            os.makedirs(outputDir)

        if csvFileName:
            out = open(csvFileName, "w", newline="", encoding="utf-8-sig")
            writer = csv.DictWriter(out, fieldnames = colNames, delimiter=",", quotechar='"', quoting=csv.QUOTE_NONNUMERIC)
            writer.writeheader()
            for row in allRows:
                rowConverted = convertRowToType(row, allTypes)
                writer.writerow(rowConverted)
            out.close()
            logger.info(f"Wrote {len(allRows)} rows to {csvFileName}.")
        
        if xlsxFileName:
            wb = Workbook()
            ws = wb.active
            ws.title = basename.split(".")[0] if "." in basename else basename
            ws.append(list(colNames))

            for row in allRows:
                rowConverted = convertRowToType(row, allTypes)
                ws.append([rowConverted[col] for col in colNames])

            wb.save(xlsxFileName)
            logger.info(f"Wrote {len(allRows)} rows to {xlsxFileName}.")

        if not csvFileName and not xlsxFileName:
            raise ValueError("Output file must have a .csv or .xlsx extension.")
        
    else:
        out = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', newline='')
        writer = csv.DictWriter(out, fieldnames = colNames, delimiter=",", quotechar='"', quoting=csv.QUOTE_NONNUMERIC)
        writer.writeheader()
        for row in allRows:
            rowConverted = convertRowToType(row, allTypes)
            writer.writerow(rowConverted)
        out.flush()
     

    
        
