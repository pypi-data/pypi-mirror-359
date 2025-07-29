
import re

from openpyxl.utils import column_index_from_string, get_column_letter

from .tokens import applyTokenReplacement

def resolveLookups(wb, elements = [], unprocessedDefinitions = [], currentElement = {}, wbMaxSize = None):
    if len(unprocessedDefinitions) == 0:
        elements.append(currentElement)
    else:
        if wbMaxSize is None:
            maxRows = 0
            maxCols = 0
            for sheet in wb.worksheets:
                maxRows = max(maxRows, sheet.max_row)
                maxCols = max(maxCols, sheet.max_column)
            wbMaxSize = (maxRows, maxCols)

        loopDefinition = unprocessedDefinitions[0]

        if not isinstance(loopDefinition, dict):
            raise ValueError(f"Invalid loop definition: {loopDefinition}")
        
        if "operation" not in loopDefinition:
            raise ValueError(f"Missing 'operation' in loop definition: {loopDefinition}")

        operation = loopDefinition["operation"].lower()
        if operation not in ["loopsheets", "findrow", "findcolumn", "findcell", "looprows", "loopcolumns"]:
            raise ValueError(f"Invalid loop operation '{operation}' in definition: {loopDefinition}")
        
        if operation == "findcell":
            if "token" in loopDefinition:
                raise ValueError(f"Cannot specify 'token' for 'findcell', use 'rowtoken' and 'columntoken' instead: {loopDefinition}")
            if "rowtoken" not in loopDefinition or "columntoken" not in loopDefinition:
                raise ValueError(f"Must specify 'rowtoken' and 'columntoken' for 'findcell': {loopDefinition}")
        else:
            if "token" not in loopDefinition:
                raise ValueError(f"Missing 'token' in loop definition: {loopDefinition}")
        
        loopElements = []
       
        if operation == "loopsheets":
            if "match" in loopDefinition:
                match = loopDefinition["match"]
                if not isinstance(match, list):
                    match = [match]
                matchingSheets = [sheet.title for sheet in wb.worksheets if sheet.title in match]
                errorStr = f"match '{match}'"
            elif "regex" in loopDefinition:
                sheetRegex = loopDefinition["regex"]
                matchingSheets = [sheet.title for sheet in wb.worksheets if re.search(sheetRegex, sheet.title)]
                errorStr = f"regex '{sheetRegex}'"
            else:
                raise ValueError(f"Missing 'match' or 'regex' in loop definition: {loopDefinition}")
            
            if len(matchingSheets) == 0:
                raise ValueError(f"No sheets matching {errorStr} in workbook: {loopDefinition}")

            if "unique" in loopDefinition and loopDefinition["unique"]:
                if len(matchingSheets) > 1:
                    raise ValueError(f"Multiple sheets matching {errorStr} in workbook: {loopDefinition}")
                
            if "select" in loopDefinition:
                if loopDefinition["select"] == "first":
                    matchingSheets = [matchingSheets[0]]
                elif loopDefinition["select"] == "last":
                    matchingSheets = [matchingSheets[-1]]
                elif loopDefinition["select"] == "all":
                    pass
                elif type(loopDefinition["select"]) == int:
                    matchingSheets = [matchingSheets[loopDefinition["select"]]]
                else:
                    raise ValueError(f"Invalid select '{loopDefinition['select']}' in definition: {loopDefinition}")

            loopElements = matchingSheets

        elif operation == "findrow" or operation == "findcolumn":
            if "match" not in loopDefinition:
                raise ValueError(f"Missing 'match' in loop definition: {loopDefinition}")
            match = loopDefinition["match"]

            if not isinstance(match, list):
                match = [match]

            if "sheet" not in loopDefinition:
                raise ValueError(f"Missing 'sheet' in loop definition: {loopDefinition}")
            sheet = applyTokenReplacement(loopDefinition["sheet"], currentElement)
            if sheet not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet}' not found in workbook: {loopDefinition}")

            if operation == "findrow":
                if "column" not in loopDefinition:
                    raise ValueError(f"Missing 'column' in loop definition: {loopDefinition}")
                searchSlice = applyTokenReplacement(loopDefinition["column"], currentElement)
            else:
                if "row" not in loopDefinition:
                    raise ValueError(f"Missing 'row' in loop definition: {loopDefinition}")
                searchSlice = applyTokenReplacement(loopDefinition["row"], currentElement)
            
            data = [str(cell.value) if cell.value is not None else "" for cell in wb[sheet][searchSlice]]
            indices = [i for i, s in enumerate(data) if s in match]

            if len(indices) == 0:
                raise ValueError(f"No matches found for '{match}' in '{sheet}' of '{currentElement['FILE_NAME']}' ({searchSlice})")

            if "unique" in loopDefinition and loopDefinition["unique"]:
                if len(indices) > 1:
                    raise ValueError(f"Multiple matches found for '{match}' in '{sheet}' of '{currentElement['FILE_NAME']}' ({searchSlice})")

            if "select" not in loopDefinition or loopDefinition["select"] == "first":
                indices = [indices[0]]
            elif loopDefinition["select"] == "last":
                indices = [indices[-1]]
            elif loopDefinition["select"] == "all":
                pass
            elif type(loopDefinition["select"]) == int:
                indices = [indices[loopDefinition["select"]]]
            else:
                raise ValueError(f"Invalid select '{loopDefinition['select']}' in definition: {loopDefinition}")

            offset = loopDefinition.get("offset", 0)
            indices = [i + offset + 1 for i in indices] # +1 to convert to 1-based index

            if operation == "findrow":
                loopElements = indices
            else:
                loopElements = [get_column_letter(i) for i in indices]

        elif operation == "findcell":
            if "match" not in loopDefinition:
                raise ValueError(f"Missing 'match' in loop definition: {loopDefinition}")
            if "sheet" not in loopDefinition:
                raise ValueError(f"Missing 'sheet' in loop definition: {loopDefinition}")
            if "select" in loopDefinition:
                raise ValueError(f"Cannot specify 'select' for 'findcell' due to ambiguity: {loopDefinition}")
            
            match = loopDefinition["match"]
            sheet = applyTokenReplacement(loopDefinition["sheet"], currentElement)

            if not isinstance(match, list):
                match = [match]

            rowToken = loopDefinition["rowtoken"]
            colToken = loopDefinition["columntoken"]            

            for col in wb[sheet].iter_cols():
                for cell in col:
                    if cell.value in match:
                        row = cell.row
                        col = get_column_letter(cell.column)

                        loopElements.append({
                            rowToken: row,
                            colToken: col
                        })

            if "unique" in loopDefinition and loopDefinition["unique"]:
                if len(loopElements) > 1:
                    raise ValueError(f"Multiple matches found for '{match}' in sheet '{sheet}'")

        elif operation == "looprows" or operation == "loopcolumns":
            if "start" not in loopDefinition:
                raise ValueError(f"Missing 'start' in loop definition: {loopDefinition}")
            start = applyTokenReplacement(loopDefinition["start"], currentElement)
            if operation == "loopcolumns":
                start = column_index_from_string(start)
            else:
                start = int(start)

            if "end" in loopDefinition and "count" in loopDefinition:
                raise ValueError("Cannot specify both 'end' and 'count' in loop definition")

            if "end" in loopDefinition:
                if loopDefinition["end"] is None:
                    end = None
                else:
                    end = applyTokenReplacement(loopDefinition["end"], currentElement)
                    if operation == "loopcolumns":
                        end = column_index_from_string(end)
                    else:
                        end = int(end)
            elif "count" in loopDefinition:
                count = applyTokenReplacement(loopDefinition["count"], currentElement)
                end = start + count - 1
            else:
                end = None

            if end is None:
                if "hint" in loopDefinition:
                    hint = applyTokenReplacement(loopDefinition["hint"], currentElement)
                    if hint not in wb.sheetnames:
                        raise ValueError(f"Sheet '{hint}' not found in workbook")
                    sheet = wb[hint]
                    if operation == "loopcolumns":
                        end = sheet.max_column
                    else:
                        end = sheet.max_row
                else:
                    if operation == "loopcolumns":
                        end = wbMaxSize[1]
                    else:
                        end = wbMaxSize[0]

            stride = loopDefinition.get("stride", 1)
            startOffset = loopDefinition.get("startoffset", 0)
            if startOffset != 0:
                start += startOffset
            endOffset = loopDefinition.get("endoffset", 0)
            if endOffset != 0:
                end += endOffset

            if end < start:
                raise ValueError(f"Start index {start} is greater than stop index {end} in definition: {loopDefinition}")

            indices = list(range(start, end + 1, stride))
            if operation == "looprows":
                loopElements = indices
            else:
                loopElements = [get_column_letter(i) for i in indices]

        if len(loopDefinition) == 0:
            raise ValueError("Loop definition is empty")
        
        for i in range(len(loopElements)):
            copy = currentElement.copy()

            if operation == "findcell":
                for key, value in loopElements[i].items():
                    if key not in copy:
                        copy[key] = value
                    else:
                        raise ValueError(f"Duplicate key '{key}' in loop definition: {loopDefinition}")
            else:
                token = loopDefinition["token"]
                if token not in copy:
                    copy[token] = loopElements[i]
                else:
                    raise ValueError(f"Duplicate token '{token}' in loop definition: {loopDefinition}")
            
            resolveLookups(wb, elements, unprocessedDefinitions[1:], copy, wbMaxSize)
