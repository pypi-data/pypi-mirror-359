"""
Core helpers para mapfeat.
* NO incluye extract_metadata (es dependiente de CARIS).
"""

from __future__ import annotations
import unicodedata
import string
from pathlib import Path
from xml.etree.ElementTree import Element, SubElement, tostring
from xml.dom import minidom
from typing import Optional, Dict, Any, Tuple, List
from lxml import etree

from openpyxl import load_workbook

# ---------- ENUM TABLES (recorta si lo deseas) ----------
REPRESENTATIVE_ATTRS = {
    "CATZOC": "Category of zone of confidence in data",
    "TECSOU": "Technique of sounding measurement",
    "STATUS": "Status",
    "CATOBS": "Category of obstruction",
    "CATPIP": "Category of pipeline/pipe",
    "CATWRK": "Category of wreck",
    "CONDTN": "Condition",
    "CONRAD": "Conspicuous, radar",
    "CONVIS": "Conspicuous, visually",
    "DUNITS": "Depth units",
    "EXPSOU": "Exposition of sounding",
    "HUNITS": "Height/length units",
    "NATCON": "Nature of construction",
    "NATSUR": "Nature of surface",
    "NATQUA": "Nature of surface – qualifying terms",
    "PRODCT": "Product",
    "QUASOU": "Quality of sounding measurement",
    "SURTYP": "Survey type",
    "VERDAT": "Vertical datum",
    "WATLEV": "Water level effect",
    "QUAPOS": "Quality of position",
    "HORDAT": "Horizontal datum",
    "cntdir": "Contour slope",
    "srffmt": "Format of bathymetric surface",
    "srfcat": "Category of bathymetric surface",
    "CoverageType": "Coverage type",
    "isotyp": "Isolation type",
    "ihmord": "Survey order",
}

# — Tabla completa de enumeraciones (par recortada aquí para brevedad) —
#   Mantengo las claves tal cual las proporcionaste; si se repite la clave,
#   Python conservará la última entrada.   
ENUM_VALUES = {
    "CATZOC": [
        ("zone of confidence A1", 1), ("zone of confidence A2", 2), ("zone of confidence B", 3),
        ("zone of confidence C", 4), ("zone of confidence D", 5),
        ("zone of confidence U (data not assessed)", 6)
    ],
    "TECSOU": [
        ("found by echo-sounder", 1), ("found by side scan sonar", 2),
        ("found by multi-beam", 3), ("found by diver", 4), ("found by lead-line", 5),
        ("swept by wire-drag", 6), ("found by laser", 7),
        ("swept by vertical acoustic system", 8), ("found by electromagnetic sensor", 9),
        ("photogrammetry", 10), ("satellite imagery", 11), ("found by levelling", 12),
        ("swept by side-scan-sonar", 13), ("computer generated", 14),
        ("found by interpherometric", 15)
    ],
    "STATUS": [
        ("permanent", 1), ("occasional", 2), ("recommended", 3), ("disused", 4),
        ("periodic/intermittent", 5), ("reserved", 6), ("temporary", 7), ("private", 8),
        ("mandatory", 9), ("{destroyed/ruined}", 10), ("extinguished", 11),
        ("illuminated", 12), ("historic", 13), ("public", 14), ("synchronized", 15),
        ("watched", 16), ("un-watched", 17), ("existence doubtful", 18),
        ("Pendiente Entrega", 19), ("Pendiente Validar", 20), ("Dato Validado", 21),
        ("Válido Para Compilación", 22), ("No Válido Para Compilación", 23),
        ("Sin procesar", 24), ("Fuente Recortada", 25), ("Fuente Combinada", 26),
        ("Dato Rescatado", 27), ("Selección Sondas Parcelario", 28)
    ],
    "CATOBS": [
        ("snag/stump", 1),
        ("wellhead", 2),
        ("diffuser", 3),
        ("crib", 4),
        ("fish haven", 5),
        ("foul area", 6),
        ("foul ground", 7),
        ("ice boom", 8),
        ("ground tackle", 9),
        ("boom", 10),
        ("fishing net", 11),
        ("well protection structure", 501),
        ("subsea installation", 502),
        ("pipeline obstruction", 503),
        ("free standing conductor pipe", 504),
        ("manifold", 505),
        ("storage tank", 506),
        ("template", 507),
        ("pontoon", 508),
        ("sundry objects", 509),
        ("Unknown", 701),
        ("Not Applicable", 703),
        ("Other", 704),
    ],
    "CATPIP": [
        ("{pipeline in general}", 1),
        ("outfall pipe", 2),
        ("intake pipe", 3),
        ("sewer", 4),
        ("bubbler system", 5),
        ("supply pipe", 6),
        ("Unknown", 701),
        ("Not Applicable", 703),
        ("Other", 704),
    ],
    "CATWRK": [
        ("non-dangerous wreck", 1),
        ("dangerous wreck", 2),
        ("distributed remains of wreck", 3),
        ("wreck showing mast/masts", 4),
        ("wreck showing any portion of hull or superstructure", 5),
        ("Unknown", 701),
        ("Not Applicable", 703),
        ("Other", 704),
    ],
    "CATZOC": [
        ("zone of confidence A1", 1),
        ("zone of confidence A2", 2),
        ("zone of confidence B", 3),
        ("zone of confidence C", 4),
        ("zone of confidence D", 5),
        ("zone of confidence U (data not assessed)", 6),
    ],
    "CONDTN": [
        ("under construction", 1),
        ("ruined", 2),
        ("under reclamation", 3),
        ("wingless", 4),
        ("planned construction", 5),
        ("operational: completed, undamaged and working normally", 501),
        ("Unknown", 701),
        ("Multiple", 702),
        ("Not Applicable", 703),
        ("Other", 704),
    ],
    "CONRAD": [
        ("radar conspicuous", 1),
        ("not radar conspicuous", 2),
        ("radar conspicuous (has radar reflector)", 3),
    ],
    "CONVIS": [
        ("visually conspicuous", 1),
        ("not visually conspicuous", 2),
    ],
    "CoverageType": [
        ("CSAR Grid (.csar)", 6),
        ("CSAR Point Cloud (.csar)", 7),
        ("CSAR Variable Resolution Surface (.csar)", 8),
    ],
    "DUNITS": [
        ("metres", 1),
        ("fathoms and feet", 2),
        ("feet", 3),
        ("fathoms and fractions", 4),
    ],
    "EXPSOU": [
        ("within the range of depth of surrounding depth area", 1),
        ("shoaler than range of depth of surrounding depth area", 2),
        ("deeper than range of depth of surrounding depth area", 3),
        ("Unknown", 701),
        ("Not Applicable", 703),
    ],
    "HORDAT": [
        ("WGS 72", 1),
        ("WGS 84", 2),
        ("European 1950 (European Datum)", 3),
        ("Potsdam datum", 4),
        ("Adindan", 5),
        ("Afgooye", 6),
        ("Ain el Abd 1970", 7),
        ("Anna 1 Astro 1965", 8),
        ("Antigua Island Astro 1943", 9),
        ("Arc 1950", 10),
        ("Arc 1960", 11),
        ("Ascension Island 1958", 12),
        ("Astro beacon `E` 1945", 13),
        ("Astro DOS 71/4", 14),
        ("Astro Tern Island FERIG) 1961", 15),
        ("Astronomical Station 1952", 16),
        ("Australian Geodetic 1966", 17),
        ("Australian Geodetic 1984", 18),
        ("Ayabelle Lighthouse", 19),
        ("Bellevue (IGN)", 20),
        ("Bermuda 1957", 21),
        ("Bissau", 22),
        ("Bogota Observatory", 23),
        ("Bukit Rimpah", 24),
        ("Camp Area Astro", 25),
        ("Campo Inchauspe", 26),
        ("Canton Astro 1966", 27),
        ("Cape", 28),
        ("Cape Canaveral", 29),
        ("Carthage", 30),
        ("Chatam Island Astro 1971", 31),
        ("Chua Astro", 32),
        ("Corrego Alegre", 33),
        ("Dabola", 34),
        ("Djakarta (Batavia)", 35),
        ("DOS 1968", 36),
        ("Easter Island 1967", 37),
        ("European 1979", 38),
        ("Fort Thomas 1955", 39),
        ("Gan 1970", 40),
        ("Geodetic Datum 1949", 41),
        ("Graciosa Base SW 1948", 42),
        ("Guam 1963", 43),
        ("Gunung Segara", 44),
        ("GUX 1 Astro", 45),
        ("Herat North", 46),
        ("Hjorsey 1955", 47),
        ("Hong Kong 1963", 48),
        ("Hu-Tzu_Shan", 49),
        ("Indian", 50),
        ("Indian 1954", 51),
        ("Indian 1975", 52),
        ("Ireland 1965", 53),
        ("ISTS 061 Astro 1968", 54),
        ("ISTS 073 Astro 1969", 55),
        ("Johnston Island 1961", 56),
        ("Kandawala", 57),
        ("Kerguelen Island 1949", 58),
        ("Kertau 1948", 59),
        ("Kusaie Astro 1951", 60),
        ("L. C. 5 Astro 1961", 61),
        ("Leigon", 62),
        ("Liberia 1964", 63),
        ("Luzon", 64),
        ("Mahe 1971", 65),
        ("Massawa", 66),
        ("Merchich", 67),
        ("Midway Astro 1961", 68),
        ("Minna", 69),
        ("Montserrat Island Astro 1958", 70),
        ("M'Poraloko", 71),
        ("Nahrwan", 72),
        ("Naparima, BWI", 73),
        ("NAD 1927 (North American 1927)", 74),
        ("NAD 1983 (North American 1983)", 75),
        ("Observatorio Meteorologico 1939", 76),
        ("Old Egyptian 1907", 77),
        ("Old Hawaiian", 78),
        ("Oman", 79),
        ("Ordnance Surrey of Great Britain 1936", 80),
        ("Pico de las Nieves", 81),
        ("Pitcairn Astro 1967", 82),
        ("Point 58", 83),
        ("Pointe Noire 1948", 84),
        ("Porto Santo 1936", 85),
        ("Provisional South American 1956", 86),
        ("Provisional South Chilean 1963 (also known as Hito XVIII 1963)", 87),
        ("Puerto Rico", 88),
        ("Quatar national", 89),
        ("Qornoq", 90),
        ("Reunion", 91),
        ("Rome 1940", 92),
        ("Santo (DOS) 1965", 93),
        ("Sao Braz", 94),
        ("Sapper Hill 1943", 95),
        ("Scharzeck", 96),
        ("Selvagem Grande 1938", 97),
        ("South American 1969", 98),
        ("South Asia", 99),
        ("Tananarive Observatory 1925", 100),
        ("Timbalai 1948", 101),
        ("Tokyo", 102),
        ("Tristan Astro 1968", 103),
        ("Viti Levu 1916", 104),
        ("Wake-Eniwetok 1960", 105),
        ("Wake Island Astro 1952", 106),
        ("Yacare", 107),
        ("Zanderij", 108),
        ("American Samoa 1962", 109),
        ("Deception Island", 110),
        ("Indian 1960", 111),
        ("Indonesian 1974", 112),
        ("North Sahara 1959", 113),
        ("Pulkovo 1942", 114),
        ("S-42 (Pulkovo 1942)", 115),
        ("S-JYSK", 116),
        ("Voirol 1950", 117),
        ("Average Terrestial System 1977", 118),
        ("Compensation Geodesique du Quebec 1977", 119),
        ("Finnish (KKJ)", 120),
        ("Ordnance Survey of Ireland", 121),
        ("Revised Kertau", 122),
        ("Revised Nahrwan", 123),
        ("GGRS 76 (Greece)", 124),
        ("Nouvelle Triangulation de France", 125),
        ("RT 90 (Sweden)", 126),
        ("Geocentric Datum of Australia (FDA)", 127),
        ("BJZ54 (A954 Beijing Coordinates)", 128),
        ("Modified BJZ54", 129),
        ("GDZ80", 130),
        ("Local datum", 131),
    ],
    "HUNITS": [
        ("metres", 1),
        ("feet", 2),
    ],
    "NATCON": [
        ("masonry", 1),
        ("concreted", 2),
        ("loose boulders", 3),
        ("hard surfaced", 4),
        ("unsurfaced", 5),
        ("wooden", 6),
        ("metal", 7),
        ("glass reinforced plastic (GRP)", 8),
        ("painted", 9),
        ("loose / unpaved", 501),
        ("loose / light", 502),
        ("hard /paved", 503),
        ("Unknown", 701),
        ("Multiple", 702),
        ("Not Applicable", 703),
        ("Other", 704),
    ],
    "NATQUA": [
        ("-", 0),
        ("fine", 1),
        ("medium", 2),
        ("coarse", 3),
        ("broken", 4),
        ("sticky", 5),
        ("soft", 6),
        ("stiff", 7),
        ("volcanic", 8),
        ("calcareous", 9),
        ("hard", 10),
        ("Unknown", 701),
        ("Multiple", 702),
        ("Not Applicable", 703),
        ("Other", 704),
    ],
    "NATSUR": [
        ("-", 0),
        ("mud", 1),
        ("clay", 2),
        ("silt/ooze", 3),
        ("sand", 4),
        ("stone", 5),
        ("gravel", 6),
        ("pebbles", 7),
        ("cobbles", 8),
        ("rock", 9),
        ("{marsh}", 10),
        ("lava", 11),
        ("{snow}", 12),
        ("{ice}", 13),
        ("coral", 14),
        ("{swamp}", 15),
        ("{bog/moor}", 16),
        ("shells", 17),
        ("boulder", 18),
        ("Unknown", 701),
        ("Multiple", 702),
        ("Not Applicable", 703),
        ("Other", 704),
    ],
    "PRODCT": [
        ("oil", 1),
        ("gas", 2),
        ("water", 3),
        ("stone", 4),
        ("coal", 5),
        ("ore", 6),
        ("chemicals", 7),
        ("drinking water", 8),
        ("milk", 9),
        ("bauxite", 10),
        ("coke", 11),
        ("iron ingots", 12),
        ("salt", 13),
        ("sand", 14),
        ("timber", 15),
        ("sawdust/wood chips", 16),
        ("scrap metal", 17),
        ("liquified natural gas (LNG)", 18),
        ("liquified petroleum gas (LPG)", 19),
        ("wine", 20),
        ("cement", 21),
        ("grain", 22),
    ],
    "QUAPOS": [
        ("surveyed", 1),
        ("unsurveyed", 2),
        ("inadequately surveyed", 3),
        ("approximate", 4),
        ("doubtful", 5),
        ("unreliable", 6),
        ("reported (not surveyed)", 7),
        ("reported (not confirmed)", 8),
        ("estimated", 9),
        ("precisely known", 10),
        ("calculated", 11),
    ],
    "QUASOU": [
        ("depth known", 1),
        ("depth unknown", 2),
        ("doubtful sounding", 3),
        ("unreliable sounding", 4),
        ("no bottom found at value shown", 5),
        ("least depth known", 6),
        ("least depth unknown, safe clearance at value shown", 7),
        ("value reported (not surveyed)", 8),
        ("value reported (not confirmed)", 9),
        ("maintained depth: the depth at which a channel is kept by human influence, usually by dredging. (IHO dictionary, S-32 5th Edition, 3057)", 10),
        ("not regularly maintained: depths may be altered by human influence, but will not be routinely maintained. (S-57 Annex A, Appendix A, IHO Object Catalogue)", 11),
        ("Not Applicable", 703),
        ("Other", 704),
    ],
    "STATUS": [
        ("permanent", 1),
        ("occasional", 2),
        ("recommended", 3),
        ("disused", 4),
        ("periodic/intermittent", 5),
        ("reserved", 6),
        ("temporary", 7),
        ("private", 8),
        ("mandatory", 9),
        ("{destroyed/ruined}", 10),
        ("extinguished", 11),
        ("illuminated", 12),
        ("historic", 13),
        ("public", 14),
        ("synchronized", 15),
        ("watched", 16),
        ("un-watched", 17),
        ("existence doubtful", 18),
        ("Pendiente Entrega", 19),
        ("Pendiente Validar", 20),
        ("Dato Validado", 21),
        ("Válido Para Compilación", 22),
        ("No Válido Para Compilación", 23),
        ("Sin procesar", 24),
        ("Fuente Recortada", 25),
        ("Fuente Combinada", 26),
        ("Dato Rescatado", 27),
        ("Selección Sondas Parcelario", 28),
    ],
    "SURTYP": [
        ("reconnaissance/sketch survey", 1),
        ("controlled survey", 2),
        ("{unsurveyed}", 3),
        ("examination survey", 4),
        ("passage survey", 5),
        ("remotely sensed", 6),
        ("INH", 7),
        ("Proyecto", 8),
        ("Compilación", 9),
    ],
    "TECSOU": [
        ("found by echo-sounder", 1),
        ("found by side scan sonar", 2),
        ("found by multi-beam", 3),
        ("found by diver", 4),
        ("found by lead-line", 5),
        ("swept by wire-drag", 6),
        ("found by laser", 7),
        ("swept by vertical acoustic system", 8),
        ("found by electromagnetic sensor", 9),
        ("photogrammetry", 10),
        ("satellite imagery", 11),
        ("found by levelling", 12),
        ("swept by side-scan-sonar", 13),
        ("computer generated", 14),
        ("found by interpherometric", 15),
    ],
    "VERDAT": [
        ("Vertical Datum is not applicable", 0),
        ("Mean Low Water Springs", 1),
        ("Mean Lower Low Water Springs", 2),
        ("Mean Sea Level", 3),
        ("Lowest Low Water", 4),
        ("Mean Low Water", 5),
        ("Lowest Low Water Springs", 6),
        ("Approximate Mean Low Water Springs", 7),
        ("Indian Spring Low Water", 8),
        ("Low Water Springs", 9),
        ("Approximate Lowest Astronomical Tide", 10),
        ("Nearly Lowest Low Water", 11),
        ("Mean Lower Low Water", 12),
        ("Low Water", 13),
        ("Approximate Mean Low Water", 14),
        ("Approximate Mean Lower Low Water", 15),
        ("Mean High Water", 16),
        ("Mean High Water Springs", 17),
        ("High Water", 18),
        ("Approximate Mean Sea Level", 19),
        ("High Water Springs", 20),
        ("Mean Higher High Water", 21),
        ("Equinoctial Spring Low Water", 22),
        ("Lowest Astronomical Tide", 23),
        ("Local Datum", 24),
        ("International Great Lakes Datum 1985", 25),
        ("Mean Water Level", 26),
        ("Lower Low Water Large Tide", 27),
        ("Higher High Water Large Tide", 28),
        ("Nearly Highest High Water", 29),
        ("Highest Astronomical Tide (HAT)", 30),
        ("Vertical Datum does not apply", 255),
        ("Mean Tide Level", 501),
        ("Unknown", 701),
        ("Not Applicable", 703),
        ("Other", 704),
    ],
    "WATLEV": [
        ("partly submerged at high water", 1),
        ("always dry", 2),
        ("always under water/submerged", 3),
        ("covers and uncovers", 4),
        ("awash", 5),
        ("subject to inundation or flooding", 6),
        ("floating", 7),
        ("Unknown", 701),
        ("Not Applicable", 703),
        ("Other", 704),
    ],
    "cntdir": [
        ("Up", 0),
        ("Down", 1),
    ],
    "cvgtyp": [
        ("Boundary", 0),
        ("Hole", 1),
    ],
    "ihmord": [
        ("Special Order", 1),
        ("Order 1A", 2),
        ("Order 1B", 3),
        ("Order 2", 4),
    ],
    "ihmown": [
        ("Autoridad Portuaria (AP)", 1),
        ("Escuela de Hidrografía (ESHIDRO)", 2),
        ("Instituto Español de Oceanografía (IEO)", 3),
        ("Instituto Hidrográfico de la Marina (IHM)", 4),
        ("Secretaría General de Pesca (SEGEPESCA)", 5),
        ("Universidad de Cádiz (UCA)", 6),
    ],
    "ihmrau": [
        ("Instituto Español de Oceanografía (IEO)", 1),
        ("Instituto Hidrográfico de La Marina (IHM)", 2),
        ("Instituto Geográfico Nacional", 3),
        ("Puertos del Estado (PE)", 4),
        ("Universidad de Cádiz (UCA)", 5),
    ],
    "ihmrre": [
        ("Superficie de Referencia Hidrográfica (SRH)", 1),
        ("G.N.S.S - RTK", 2),
        ("Mareógrafo múltiple (ZDF)", 3),
        ("Mareógrafo único", 4),
        ("Predicción", 5),
        ("Regla de mareas", 6),
    ],
    "ihmrty": [
        ("Permanente", 1),
        ("Temporal", 2),
    ],
    "ihmsou": [
        ("Atlas Deso10", 1),
        ("Atlas Deso20", 2),
        ("Atlas Deso25", 3),
        ("Kongsberg EA400", 4),
        ("Kongsberg EA600", 5),
        ("Escandallo", 6),
        ("Atlas Hydrosweep DS-3", 7),
        ("Kongsberg Geoswath 250", 8),
        ("Kongsberg Geoswath 500", 9),
        ("Kongsberg EM12", 10),
        ("Kongsberg EM120", 11),
        ("Kongsberg EM1000", 12),
        ("Kongsberg EM1002", 13),
        ("Kongsberg EM2040", 14),
        ("Kongsberg EM2040C", 15),
        ("Kongsberg EM300", 16),
        ("Kongsberg EM302", 17),
        ("Kongsberg EM3000", 18),
        ("Kongsberg EM3000D", 19),
        ("Kongsberg EM3002", 20),
        ("Kongsberg EM3002D", 21),
        ("Kongsberg EM710", 22),
        ("Kongsberg EM712", 23),
        ("ELAC SEABEAM 1184", 24),
        ("Teledyne Reson T20P", 25),
        ("Kongsberg EM2040P", 26),
        ("Kongsberg EM122", 27),
    ],
    "ihmspo": [
        ("Rhothetha", 1),
        ("Trisponder", 2),
        ("G.P.S. Autónomo", 3),
        ("G.P.S. Diferencial Egnos", 4),
        ("G.P.S. Diferencial Estación Base", 5),
        ("G.P.S. Diferencial Omnistar", 6),
        ("G.P.S. Diferencial Omnistart XP", 7),
        ("G.P.S. Diferencial RTK", 8),
    ],
    "ihmves": [
        ("Antares", 1),
        ("Astrolabio", 2),
        ("Cástor", 3),
        ("Escandallo", 4),
        ("Escuela de Hidrografía", 5),
        ("Hespérides", 6),
        ("Malaspina", 7),
        ("Póllux", 8),
        ("Rigel", 9),
        ("Sondaleza", 10),
        ("Tofiño", 11),
        ("Ángeles Alvariño", 12),
        ("Emma Bardán", 13),
        ("Miguel Oliver", 14),
        ("Ramón Margalef", 15),
        ("Sarmiento de Gamboa", 16),
        ("UCádiz", 17),
        ("Vizconde de Eza", 18),
        ("Zodiac", 19),
        ("Autoridad Portuaria (AP)", 20),
        ("Narwhal", 21),
    ],
    "isotyp": [
        ("Shoal", 0),
        ("Deep", 1),
    ],
    "srfcat": [
        ("Standard", 0),
        ("Product", 1),
    ],
    "srffmt": [
        ("CARIS HIPS", 0),
    ],
}
# ---------------------------------------------------------

def normalize_text(text: str) -> str:
    """ Convierte ñ→n y quita tildes/acentos. """
    text = text.replace("ñ", "n").replace("Ñ", "N")
    text = unicodedata.normalize("NFKD", text)
    return "".join(c for c in text if not unicodedata.combining(c))

# -------- utilidades Excel --------------------------------
def letters_range(start: str, end: str) -> List[str]:
    """Devuelve lista de columnas Excel entre start y end incluidas."""
    alphabet = list(string.ascii_uppercase)

    def col_to_num(col: str) -> int:
        num = 0
        for c in col:
            num = num * 26 + alphabet.index(c.upper()) + 1
        return num

    def num_to_col(num: int) -> str:
        col = ""
        while num:
            num, rem = divmod(num - 1, 26)
            col = alphabet[rem] + col
        return col

    return [num_to_col(n) for n in range(col_to_num(start), col_to_num(end) + 1)]


COLS = (
    letters_range("B", "B")
    + letters_range("D", "D")
    + letters_range("F", "G")
    + letters_range("O", "Z")
    + letters_range("AD", "AN")
    + letters_range("AU", "AX")
)
HEADERS_ROW = 8
DATA_ROW = 9
OBJNAM_COL = "F"

# ----------- mapeo enum -----------------------------------
def map_enum(attr: str, value: Any) -> Any:
    """Sustituye texto por código numérico si existe en ENUM_VALUES."""
    if not isinstance(value, str):
        return value
    options = ENUM_VALUES.get(attr)
    if not options:
        return value
    lookup = value.strip().casefold()
    for text, code in options:
        if lookup == text.casefold():
            return code
    return value

# ---------- Excel → dict ----------------------------------
def read_excel(xlsx_path: Path) -> Tuple[str, Dict[str, Any]]:
    """Lee la fila 8/9 y devuelve (OBJNAM, dict de metadatos)."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    headers = [ws[f"{c}{HEADERS_ROW}"].value for c in COLS]
    values  = [ws[f"{c}{DATA_ROW}"].value for c in COLS]
    data = dict(zip([h.strip() if h else "" for h in headers], values))

    objnam_value = ws[f"{OBJNAM_COL}{DATA_ROW}"].value
    if not objnam_value:
        raise ValueError(f"OBJNAM vacío en {xlsx_path}")
    return str(objnam_value), data

def remove_blank_lines_from_file(file_path: Path):
    """Elimina líneas completamente vacías o con solo espacios de un archivo."""
    lines = file_path.read_text(encoding="utf-8").splitlines()
    non_blank = [line for line in lines if line.strip()]
    file_path.write_text("\n".join(non_blank) + "\n", encoding="utf-8")	
	
# ------------ Generador XML -------------------------------
def build_xml2(
    data: Dict[str, Any],
    extra: Optional[Dict[str, Any]] = None,
    indent: str = "  ",
) -> bytes:
    """Funde `data` y `extra`, normaliza textos y devuelve XML bytes."""
    if extra:
        for k, v in extra.items():
            if k not in data or data[k] in (None, ""):
                data[k] = v

    root = Element("Attributes")
    for tag, val in data.items():
        if not tag or val in (None, ""):
            continue
        tag_clean = tag.strip()
        if not tag_clean.replace("_", "").isalnum():
            raise ValueError(f"Etiqueta XML no válida: {tag_clean}")
        el = SubElement(root, tag_clean)
        val_final = normalize_text(str(map_enum(tag_clean, val)))
        el.text = val_final

    xml_bytes = tostring(root, encoding="utf-8")
    pretty = minidom.parseString(xml_bytes).toprettyxml(indent=indent, encoding="UTF-8")

    filtered = b"\n".join([line for line in pretty.splitlines() if line.strip()])
    return filtered

def build_xml(
    data: Dict[str, Any],
    extra: Optional[Dict[str, Any]] = None,
    indent: str = "  "
) -> bytes:
    """
    Funde `data` y `extra`, normaliza textos, aplica map_enum y genera XML con lxml.
    Devuelve XML como bytes (con declaración y pretty print).
    """
    if extra:
        for k, v in extra.items():
            if k not in data or data[k] in (None, ""):
                data[k] = v

    root = etree.Element("Attributes")

    for tag, val in data.items():
        if not tag or val in (None, ""):
            continue
        tag_clean = tag.strip()
        if not tag_clean.replace("_", "").isalnum():
            raise ValueError(f"Etiqueta XML no válida: {tag_clean}")
        el = etree.SubElement(root, tag_clean)
        val_final = normalize_text(str(map_enum(tag_clean, val)))
        el.text = val_final

    return etree.tostring(
        root,
        pretty_print=True,
        encoding="UTF-8",
        xml_declaration=True
    )