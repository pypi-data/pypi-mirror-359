import os
import pytest
import openpyxl
from meyigi_scripts.file_writers.append_to_excel import append_to_excel



@pytest.fixture
def temp_excel_file():
    """Creates a temporary Excel file for testing."""
    filename = "data/test_output.xlsx"
    if os.path.exists(filename):
        os.remove(filename)
    yield filename
    if os.path.exists(filename):
        os.remove

@pytest.mark.excel
def test_append_to_excel_creates_file(temp_excel_file):
    """Tests that append_to_excel creates an Excel file."""
    data = {"Name" : "Meyigi", "Age" : 34, "city" : "Naryn"}

    append_to_excel(data, temp_excel_file)

    assert os.path.exists(temp_excel_file) # Check if the file was created

@pytest.mark.excel
def test_append_to_excel_writed_data(temp_excel_file):
    data = {"Name" : "Meyigi", "Age" : 34, "city" : "Naryn"}

    append_to_excel(data, temp_excel_file)

    workbook = openpyxl.load_workbook(temp_excel_file)
    sheet = workbook.active

    assert sheet.cell(row=1, column=1).value == "Name"
    assert sheet.cell(row=1, column=2).value == "Age"
    assert sheet.cell(row=1, column=3).value == "city"
    assert sheet.cell(row=2, column=1).value == "Meyigi"
    assert sheet.cell(row=2, column=2).value == 34
    assert sheet.cell(row=2, column=3).value == "Naryn"

@pytest.mark.excel
def test_append_to_excel_writed_data(temp_excel_file):
    data = [
        {"Name" : "Meyigi", "Age" : 34, "city" : "Naryn"},
        {"Name" : "Daniel", "Age" : 24, "city" : "New-York"},
    ]

    append_to_excel(data, temp_excel_file)

    workbook = openpyxl.load_workbook(temp_excel_file)
    sheet = workbook.active

    assert sheet.cell(row=1, column=1).value == "Name"
    assert sheet.cell(row=1, column=2).value == "Age"
    assert sheet.cell(row=1, column=3).value == "city"
    assert sheet.cell(row=2, column=1).value == "Meyigi"
    assert sheet.cell(row=2, column=2).value == 34
    assert sheet.cell(row=2, column=3).value == "Naryn"
    assert sheet.cell(row=3, column=1).value == "Daniel"
    assert sheet.cell(row=3, column=2).value == 24
    assert sheet.cell(row=3, column=3).value == "New-York"

@pytest.mark.excel
def test_append_to_excel_raises_type_error(temp_excel_file):
    """Test that append_to_excel raises a TypeError if input is not a dict."""
    with pytest.raises(TypeError):
        append_to_excel(["not", "a", "dict"], temp_excel_file)