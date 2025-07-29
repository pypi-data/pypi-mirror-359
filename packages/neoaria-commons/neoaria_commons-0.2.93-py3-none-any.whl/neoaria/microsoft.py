import openpyxl

class Excel:

    #FIXME 여기 하드코딩함.
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)
        self.active_work_sheet = self.wb.active

    def getWorkbook(self):
        return self.workbook

    def getWorkSheet(self, sheet_name:str)->openpyxl.worksheet.worksheet.Worksheet:
        return self.workbook[sheet_name]

class PowerPoint:

    #TODO 여기도 하드코딩함.
    def __init__(self, file_path):
        self.file_path = file_path
        self.presentation = openpyxl.load_workbook(file_path)
        self.active_slide = self.presentation.active

    def createSlide(self):
        pass
